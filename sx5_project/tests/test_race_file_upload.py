import io

import openpyxl
import pandas as pd
from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase, override_settings
from django.contrib.auth.models import User


from classifications.models import Classification, ClassificationResult
from races.models import Race, Result, Runner, Season


def generate_file_with_runners():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(
        ["First Name", "Last Name", "Participant Number", "Category", "Club", "Time"]
    )
    ws.append(["Lukasz", "Pudlo", "121", "MS", "Unaffiliated", "00:19:31"])
    ws.append(["Callum", "Wallace", "654", "MS",
              "Bellahouston Harriers", "00:20:51"])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    uploaded_file = SimpleUploadedFile(
        "test_kings.xlsx",
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return uploaded_file


@override_settings(
    STORAGES={
        "default": {
            "BACKEND": "django.core.files.storage.FileSystemStorage",
        },
        "staticfiles": {
            "BACKEND": "django.contrib.staticfiles.storage.StaticFilesStorage",
        },
    }
)
class TestFileUpload(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.season = Season.objects.create(season_start_year=2024)
        cls.season = Season.objects.create(season_start_year=2025)

        from django.contrib.auth.models import User

        cls.user = User.objects.create_user(
            username="testuser", password="testpass123")

    def setUp(self):
        self.client.login(username="testuser", password="testpass123")

    def test_race_file_upload_creates_race(self):
        uploaded_file = generate_file_with_runners()

        form_data = {
            "name": "Test King's Park",
            "season_start_year": 2025,
            "race_file": uploaded_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 1)

        # Verify the race has the correct data
        race = Race.objects.first()
        self.assertEqual(race.name, "Test King's Park")
        self.assertEqual(race.season_start_year, 2025)

    def test_runners_are_created_for_each_season(self):
        # Create King's Park race for season 2024/2025
        uploaded_file = generate_file_with_runners()

        form_data = {
            "name": "Test King's Park",
            "season_start_year": 2024,
            "race_file": uploaded_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Create King's Park race for season 2025/2026
        wb_kings_2025 = openpyxl.Workbook()
        ws = wb_kings_2025.active

        ws.append(
            [
                "First Name",
                "Last Name",
                "Participant Number",
                "Category",
                "Club",
                "Time",
            ]
        )
        ws.append(["Lukasz", "Pudlo", "121", "MS", "Unaffiliated", "00:19:31"])
        ws.append(
            ["Callum", "Wallace", "987", "MS", "Bellahouston Harriers", "00:20:51"]
        )

        excel_file = io.BytesIO()
        wb_kings_2025.save(excel_file)
        excel_file.seek(0)

        uploaded_file = SimpleUploadedFile(
            "test_2025_kings.xlsx",
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        form_data = {
            "name": "Test King's Park",
            "season_start_year": 2025,
            "race_file": uploaded_file,
        }

        # Post to the correct URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Check that there are two races
        race_count = Race.objects.all().count()
        self.assertEqual(race_count, 2, "The race count is not 2")

        runner_count = Runner.objects.all().count()

        self.assertEqual(
            runner_count, 4, f"Expected 4 runners, got {runner_count}")

        # Get the specific Callum runner for season 2025
        season_2025 = Season.objects.get(season_start_year=2025)
        runner_callum_2025 = Runner.objects.get(
            first_name="Callum", last_name="Wallace", season=season_2025
        )

        self.assertEqual(
            runner_callum_2025.season.season_start_year,
            2025,
            "Callum runner object's season is not 2025",
        )

        season_2024 = Season.objects.get(season_start_year=2024)
        runner_callum_2024 = Runner.objects.get(
            first_name="Callum", last_name="Wallace", season=season_2024
        )
        self.assertEqual(
            runner_callum_2024.season.season_start_year,
            2024,
            "Callum runner object's season is not 2024",
        )

        self.assertEqual(
            runner_callum_2024.participant_number,
            "654",
            "Callum's participant number for season 2024/2025 is not correct",
        )

    def test_numbers_with_hyphens_are_allowed(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(
            [
                "First Name",
                "Last Name",
                "Participant Number",
                "Category",
                "Club",
                "Time",
            ]
        )
        ws.append(["Lukasz", "Pudlo", "121-963",
                  "MS", "Unaffiliated", "00:19:31"])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        uploaded_file = SimpleUploadedFile(
            "test_kings.xlsx",
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        form_data = {
            "name": "Test King's Park",
            "season_start_year": 2025,
            "race_file": uploaded_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 1)

        season_2025 = Season.objects.get(season_start_year=2025)
        runner = Runner.objects.get(
            first_name="Lukasz", last_name="Pudlo", season=season_2025
        )

        self.assertEqual(
            runner.participant_number,
            "121-963",
            f"The runner's participant number is not correct. Got: '{runner.participant_number}'",
        )

    def test_file_is_read(self):
        file_path = f"{settings.BASE_DIR}/tests/files/2025/kings.xlsx"

        with open(file_path, "rb") as f:
            uploaded_file = SimpleUploadedFile(
                "kings.xlsx",
                f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        form_data = {
            "name": "Test King's Park",
            "season_start_year": 2025,
            "race_file": uploaded_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 1)

        # Check that the number of runners in the file
        # is the same as the number of runners created
        df = pd.read_excel(file_path)
        row_count = df["First Name"].count()
        runner_count = Runner.objects.count()

        self.assertEqual(
            row_count,
            runner_count,
            "The number of rows is not equal to the number of runners created.",
        )


class TestFieldCorrectness(TestCase):
    maxDiff = None

    @classmethod
    def setUpTestData(cls):
        cls.season = Season.objects.create(season_start_year=2024)
        cls.season = Season.objects.create(season_start_year=2025)
        cls.user = User.objects.create_user(
            username="testuser", password="testpass123")

    def setUp(self):
        self.client.login(username="testuser", password="testpass123")

        # Create a race based on true result data
        file_path = f"{settings.BASE_DIR}/tests/files/2025/kings.xlsx"
        with open(file_path, "rb") as f:
            uploaded_file = SimpleUploadedFile(
                "kings.xlsx",
                f.read(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

        form_data = {
            "name": "Test King's Park",
            "season_start_year": 2025,
            "race_file": uploaded_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 1)

    # Test that the values are the same in the file and in the result table
    def test_values_are_the_same(self):
        file_path = f"{settings.BASE_DIR}/tests/files/2025/kings.xlsx"

        df = pd.read_excel(file_path, dtype=str)
        df = df.fillna("")
        first_name_list_df = df["First Name"].values.tolist()
        last_name_list_df = df["Last Name"].values.tolist()
        participant_number_list_df = df["Participant Number"].values.tolist()
        category_list_df = df["Category"].values.tolist()
        time_list_df = df["Time"].values.tolist()

        # Replace empty strings in the Excel with "Unaffiliated"
        df.loc[df["Club"] == "", "Club"] = "Unaffiliated"
        club_list_df = df["Club"].values.tolist()

        first_name_list_app = []
        last_name_list_app = []
        participant_number_list_app = []
        category_list_app = []
        club_list_app = []
        time_list_app = []

        results = Result.objects.filter(race=1)
        for result in results:
            first_name_list_app.append(result.runner.first_name)
            last_name_list_app.append(result.runner.last_name)
            participant_number_list_app.append(
                result.runner.participant_number)
            category_list_app.append(result.runner.category)
            club_list_app.append(result.runner.club)
            time_list_app.append(result.time)

        self.assertEqual(
            first_name_list_df, first_name_list_app, "The " "first names are different"
        )
        self.assertEqual(
            last_name_list_df, last_name_list_app, "The " "last names are different"
        )
        self.assertEqual(
            participant_number_list_df,
            participant_number_list_app,
            "The " "participant numbers " "names are different",
        )
        self.assertEqual(
            category_list_df, category_list_app, "The " "category are different"
        )
        self.assertEqual(club_list_df, club_list_app,
                         "The " "clubs are different")
        self.assertEqual(
            time_list_df, time_list_app, "The " "time result are different"
        )


def generate_kings():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(
        ["First Name", "Last Name", "Participant Number", "Category", "Club", "Time"]
    )
    ws.append(["Andrew", "Anderson", "727", "M40",
              "Cambuslang Harriers", "00:18:59"])
    ws.append(["Callum", "Wallace", "811", "MS",
              "Bellahouston Harriers", "00:19:07"])
    ws.append(["Ethan", "Tyler", "814", "MS",
              "Garscube Harriers", "00:19:15"])
    ws.append(["Richard", "Cooper", "734", "MS",
              "Cambuslang Harriers", "00:19:42"])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    uploaded_file = SimpleUploadedFile(
        "test_kings.xlsx",
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return uploaded_file


def generate_linn():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(
        ["First Name", "Last Name", "Participant Number", "Category", "Club", "Time"]
    )
    ws.append(["Finlay", "Murray", "892", "MS",
              "East Sutherland Athletics Club", "00:23:39"])
    ws.append(["Ethan", "Tyler", "814", "MS",
              "Garscube Harriers", "00:24:36"])
    ws.append(["Callum", "Wallace", "811", "MS",
              "Bellahouston Harriers", "00:24:44"])
    ws.append(["Richard", "Cooper", "734", "MS",
              "Cambuslang Harriers", "00:25:15"])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    uploaded_file = SimpleUploadedFile(
        "test_linn.xlsx",
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return uploaded_file


def generate_rouken():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(
        ["First Name", "Last Name", "Participant Number", "Category", "Club", "Time"]
    )
    ws.append(["Andrew", "Anderson", "727", "M40",
              "Cambuslang Harriers", "00:21:00"])
    ws.append(["Callum", "Wallace", "811", "MS",
              "Bellahouston Harriers", "00:21:03"])
    ws.append(["Richard", "Cooper", "734", "MS",
              "Cambuslang Harriers", "00:21:28"])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    uploaded_file = SimpleUploadedFile(
        "test_rouken.xlsx",
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return uploaded_file


def generate_pollok():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(
        ["First Name", "Last Name", "Participant Number", "Category", "Club", "Time"]
    )
    ws.append(["Andrew", "Anderson", "727", "M40",
              "Cambuslang Harriers", "00:21:00"])
    ws.append(["Callum", "Wallace", "811", "MS",
              "Bellahouston Harriers", "00:21:03"])
    ws.append(["Richard", "Cooper", "734", "MS",
              "Cambuslang Harriers", "00:21:28"])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    uploaded_file = SimpleUploadedFile(
        "test_rouken.xlsx",
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    return uploaded_file


class TestClassificationRecalculation(TestCase):
    maxDiff = None

    @classmethod
    def setUpTestData(cls):
        cls.season = Season.objects.create(season_start_year=2025)
        cls.user = User.objects.create_user(
            username="testuser", password="testpass123")

    def setUp(self):
        self.client.login(username="testuser", password="testpass123")

        # Add King's Park race
        kings_file = generate_kings()

        form_data = {
            "name": "KP Classification Recalculation Test",
            "season_start_year": 2025,
            "race_file": kings_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 1)

        # Add Linn Park race
        linn_file = generate_linn()

        form_data = {
            "name": "LP Classification Recalculation Test",
            "season_start_year": 2025,
            "race_file": linn_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 2)

        # Add Rouken Glen race
        rouken_file = generate_rouken()

        form_data = {
            "name": "RG Classification Recalculation Test",
            "season_start_year": 2025,
            "race_file": rouken_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 3)

        # Add Pollok Park race
        pollok_file = generate_pollok()

        form_data = {
            "name": "PP Classification Recalculation Test",
            "season_start_year": 2025,
            "race_file": pollok_file,
        }

        # Post to the URL: /races/race/new/
        response = self.client.post("/races/race/new/", form_data)

        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 4)

    def test_nr_of_points_is_correct(self):
        # Get the classification after Rouken Glen race
        classification_after_rg = Classification.objects.get(
            slug="kp-classification-recalculation-test-2025-classification")
        callum_runner_object = Runner.objects.get(
            first_name="Callum", last_name="Wallace")
        callum_classification_result_after_rg = ClassificationResult.objects.get(
            runner=callum_runner_object, classification=classification_after_rg)

        overall_nr_of_races = Race.objects.all().count()
        self.assertEqual(overall_nr_of_races, 3,
                         "The number of races is not 3")

        # Check that Callum Wallace is in the classification results after Rouken Glen
        self.assertEqual(
            callum_classification_result_after_rg.runner.participant_number, "811", "Lukasz's participant number is incorrect")

        self.assertEqual(callum_classification_result_after_rg.general_points,
                         4, "Callum Wallace has an incorrect number of points after Rouken Glen")

    def test_runner_out_of_series(self):
        pollok_classification = Classification.objects.get(
            slug="pp-classification-recalculation-test-2025-classification")
        pollok_classification_runner_count = ClassificationResult.objects.filter(
            classification=pollok_classification).count()
        self.assertEqual(pollok_classification_runner_count, 3,
                         "There should be only 3 runners in the classification after Pollok Park")
        # Check that the correct runner was removed from the classification
        callum_runner_object = Runner.objects.get(participant_number="811")
        andrew_runner_object = Runner.objects.get(participant_number="727")
        richard_runner_object = Runner.objects.get(participant_number="734")

        self.assertEqual(callum_runner_object.participant_number, "811",
                         "Callum's number is incorrect")
        self.assertEqual(andrew_runner_object.participant_number, "727",
                         "Andrew's number is incorrect")
        self.assertEqual(richard_runner_object.participant_number, "734",
                         "Richard's number is incorrect")
