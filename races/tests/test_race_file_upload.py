from django.test import TestCase
from django.core.files.uploadedfile import SimpleUploadedFile
import io
import openpyxl

from races.models import Race, Result, Runner, Season


def generate_file_with_runners():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws.append(['First Name', 'Last Name', 'Participant Number',
               'Category', 'Club', 'Time'])
    ws.append(['Lukasz', 'Pudlo', '121', 'MS', 'Unaffiliated', '00:19:31'])
    ws.append(['Callum', 'Wallace', '654', 'MS',
               'Bellahouston Harriers', '00:20:51'])

    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    uploaded_file = SimpleUploadedFile(
        'test_kings.xlsx',
        excel_file.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    return uploaded_file


class TestFileUpload(TestCase):

    @classmethod
    def setUpTestData(cls):
        cls.season = Season.objects.create(season_start_year=2024)
        cls.season = Season.objects.create(season_start_year=2025)

        from django.contrib.auth.models import User
        cls.user = User.objects.create_user(
            username='testuser',
            password='testpass123'
        )

    def setUp(self):
        self.client.login(username='testuser', password='testpass123')

    def test_race_file_upload_creates_race(self):
        uploaded_file = generate_file_with_runners()

        form_data = {
            'name': "Test King's Park",
            'season_start_year': 2025,
            'race_file': uploaded_file
        }

        # Post to the URL: /races/race/new/
        response = self.client.post('/races/race/new/', form_data)

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 1)

        # Verify the race has the correct data
        race = Race.objects.first()
        self.assertEqual(race.name, "Test King's Park")
        self.assertEqual(race.season_start_year, 2025)

    def test_runners_are_created_for_each_season(self):
        # Create King's Park race for season 2024/2025
        uploaded_file = generate_file_with_runners()

        form_data = {
            'name': "Test King's Park",
            'season_start_year': 2024,
            'race_file': uploaded_file
        }

        # Post to the URL: /races/race/new/
        response = self.client.post('/races/race/new/', form_data)

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Create King's Park race for season 2025/2026
        wb_kings_2025 = openpyxl.Workbook()
        ws = wb_kings_2025.active

        ws.append(['First Name', 'Last Name', 'Participant Number',
                   'Category', 'Club', 'Time'])
        ws.append(['Lukasz', 'Pudlo', '121', 'MS', 'Unaffiliated', '00:19:31'])
        ws.append(['Callum', 'Wallace', '987', 'MS',
                   'Bellahouston Harriers', '00:20:51'])

        excel_file = io.BytesIO()
        wb_kings_2025.save(excel_file)
        excel_file.seek(0)

        uploaded_file = SimpleUploadedFile(
            'test_2025_kings.xlsx',
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        form_data = {
            'name': "Test King's Park",
            'season_start_year': 2025,
            'race_file': uploaded_file
        }

        # Post to the correct URL: /races/race/new/
        response = self.client.post('/races/race/new/', form_data)

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Check that there are two races
        race_count = Race.objects.all().count()
        self.assertEqual(race_count, 2, "The race count is not 2")

        runner_count = Runner.objects.all().count()

        self.assertEqual(
            runner_count, 4, f"Expected 4 runners, got {runner_count}")

        # Get the specific Callum runner for season 2025
        season_2025 = Season.objects.get(season_start_year=2025)
        runner_callum_2025 = Runner.objects.get(
            first_name='Callum',
            last_name='Wallace',
            season=season_2025
        )

        self.assertEqual(runner_callum_2025.season.season_start_year, 2025,
                         "Callum runner object's season is not 2025")

        season_2024 = Season.objects.get(season_start_year=2024)
        runner_callum_2024 = Runner.objects.get(
            first_name='Callum',
            last_name='Wallace',
            season=season_2024
        )
        self.assertEqual(runner_callum_2024.season.season_start_year, 2024,
                         "Callum runner object's season is not 2024")

        self.assertEqual(runner_callum_2024.participant_number, "654",
                         "Callum's participant number for season 2024/2025 is not correct")

    def test_numbers_with_hyphens_are_allowed(self):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws.append(['First Name', 'Last Name', 'Participant Number',
                   'Category', 'Club', 'Time'])
        ws.append(['Lukasz', 'Pudlo', '121-963',
                  'MS', 'Unaffiliated', '00:19:31'])

        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        uploaded_file = SimpleUploadedFile(
            'test_kings.xlsx',
            excel_file.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

        form_data = {
            'name': "Test King's Park",
            'season_start_year': 2025,
            'race_file': uploaded_file
        }

        # Post to the URL: /races/race/new/
        response = self.client.post('/races/race/new/', form_data)

        for runner in Runner.objects.all():
            print(
                f"  - {runner.first_name} {runner.last_name} (Participant Number: {runner.participant_number})")

        # Check that we got a redirect (302 status code means success)
        self.assertEqual(response.status_code, 302)

        # Check that the race was created
        self.assertEqual(Race.objects.count(), 1)

        season_2025 = Season.objects.get(season_start_year=2025)
        runner = Runner.objects.get(
            first_name="Lukasz",
            last_name="Pudlo",
            season=season_2025
        )

        self.assertEqual(runner.participant_number, '121-963',
                         f"The runner's participant number is not correct. Got: '{runner.participant_number}'")
